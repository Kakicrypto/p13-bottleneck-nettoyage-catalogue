"""
rapport_anomalies.py
Génère le rapport d'anomalies exploitable pour le service achats (Décision 9),
avec priorisation (Décision 12) : 🔴 Critique / 🟠 Majeure / 🟡 Mineure.

Ce script réutilise les schémas Pandera déjà validés dans test_pandera.py et
ajoute une couche de MAPPING des erreurs vers une priorité métier, puis exporte
un fichier Excel lisible par une personne non-technique (Décision 9 du CDC).
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import pandera.pandas as pa
from test_pandera import (
    schema_erp,
    schema_web,
    schema_liaison,
    schema_complet,
    verifier_correspondance_erp_liaison,
)
from data_prep import charger_donnees_brutes, nettoyer_web, construire_df_complet


# ---------------------------------------------------------------------------
# Décision 12 : mapping règle -> priorité métier
# Clé = le texte de l'erreur tel que défini dans les schémas Pandera (test_pandera.py)
# ---------------------------------------------------------------------------
PRIORITE_PAR_REGLE = {
    "greater_than_or_equal_to(0)": "🟠 Majeure",  # défaut par colonne, affiné plus bas
    "Incohérence entre stock_status et stock_quantity": "🟠 Majeure",
    "Marge négative détectée : price <= purchase_price": "🔴 Critique",
    "SKU au format non numérique détecté": "🟡 Mineure",
    "column 'sku' not in dataframe": "🟡 Mineure",
    "not_nullable": "🟠 Majeure",
    "field_uniqueness": "🟠 Majeure",
}

# Cas particuliers : le même check "ge(0)" a une priorité différente selon la colonne
PRIORITE_PAR_COLONNE = {
    "price": "🔴 Critique",           # prix négatif = vente à perte immédiate
    "stock_quantity": "🔴 Critique",  # stock négatif = rupture mal signalée
    "purchase_price": "🟠 Majeure",   # fausse la marge mais pas de décision immédiate bloquée
}


def determiner_priorite(row):
    """Applique la logique de priorisation définie en Décision 12."""
    colonne = row.get("column")
    check = str(row.get("check", ""))

    if colonne in PRIORITE_PAR_COLONNE and "greater_than_or_equal_to" in check:
        return PRIORITE_PAR_COLONNE[colonne]

    for cle, priorite in PRIORITE_PAR_REGLE.items():
        if cle in check:
            return priorite

    return "🟡 Mineure"  # valeur par défaut si règle non cartographiée


ACTION_RECOMMANDEE = {
    "🔴 Critique": "Corriger aujourd'hui : vérifier la saisie avec le fournisseur/l'ERP",
    "🟠 Majeure": "Corriger cette semaine : fiabilise l'analyse, pas d'urgence commerciale",
    "🟡 Mineure": "Corriger quand possible : défaut structurel sans impact business direct",
}


def detail_marge_negative(df_erp):
    """Détaille ligne par ligne les cas où price <= purchase_price (marge négative)."""
    mask = df_erp["price"] <= df_erp["purchase_price"]
    sous = df_erp[mask]
    return pd.DataFrame(
        {
            "source": "erp",
            "index": sous.index,
            "column": "price / purchase_price",
            "check": "Marge négative détectée : price <= purchase_price",
            "failure_case": [
                f"price={p} <= purchase_price={pp}"
                for p, pp in zip(sous["price"], sous["purchase_price"])
            ],
        }
    )


def detail_coherence_stock(df_erp):
    """Détaille ligne par ligne les incohérences stock_status / stock_quantity."""
    coherent = (df_erp["stock_quantity"] <= 0) == (df_erp["stock_status"] == "outofstock")
    sous = df_erp[~coherent]
    return pd.DataFrame(
        {
            "source": "erp",
            "index": sous.index,
            "column": "stock_status / stock_quantity",
            "check": "Incohérence entre stock_status et stock_quantity",
            "failure_case": [
                f"stock_quantity={q}, stock_status={s}"
                for q, s in zip(sous["stock_quantity"], sous["stock_status"])
            ],
        }
    )


def detail_sku_non_numerique(df_web_clean):
    """Détaille ligne par ligne les sku au format non numérique."""
    numerique = df_web_clean["sku"].astype(str).str.isnumeric()
    sous = df_web_clean[~numerique]
    return pd.DataFrame(
        {
            "source": "web",
            "index": sous.index,
            "column": "sku",
            "check": "SKU au format non numérique détecté",
            "failure_case": sous["sku"].astype(str),
        }
    )


def collecter_erreurs(schema, df, nom_source):
    """Valide un dataframe et retourne un DataFrame d'anomalies structuré.

    NB : les règles portant sur une seule colonne (Column Check) donnent déjà
    l'index exact de chaque ligne fautive via Pandera. En revanche, les règles
    "dataframe-level" (comparaison entre 2 colonnes, cohérence croisée) ne
    remontent qu'un seul résultat agrégé (vrai/faux) sans indiquer QUELLES
    lignes posent problème. On règle ce cas avec les fonctions detail_* ci-dessus,
    qui recalculent le masque booléen pour lister chaque ligne fautive.
    """
    try:
        schema.validate(df, lazy=True)
        return pd.DataFrame()  # aucune anomalie
    except pa.errors.SchemaErrors as e:
        erreurs = e.failure_cases.copy()
        # on retire les lignes "dataframe-level" agrégées (index NaN) : on les
        # remplace par le détail ligne par ligne calculé séparément
        erreurs_colonnes = erreurs[erreurs["column"].notna()].copy()
        erreurs_colonnes["source"] = nom_source
        erreurs_colonnes["priorite"] = erreurs_colonnes.apply(determiner_priorite, axis=1)
        erreurs_colonnes["action_recommandee"] = erreurs_colonnes["priorite"].map(ACTION_RECOMMANDEE)
        return erreurs_colonnes[
            ["source", "index", "column", "check", "failure_case", "priorite", "action_recommandee"]
        ]


def generer_rapport():
    df_erp, df_web, df_liaison = charger_donnees_brutes()
    df_web_clean = nettoyer_web(df_web)
    df_complet = construire_df_complet(df_erp, df_web, df_liaison)

    # correctif type mixte déjà identifié lors du test comparatif (Décision 6)
    df_web_clean = df_web_clean.copy()

    morceaux = [
        collecter_erreurs(schema_erp, df_erp, "erp"),
        collecter_erreurs(schema_web, df_web_clean, "web"),
        collecter_erreurs(schema_liaison, df_liaison, "liaison"),
        collecter_erreurs(schema_complet, df_complet, "complet (post-fusion)"),
    ]

    # Détail ligne par ligne des règles multi-colonnes (voir docstring collecter_erreurs)
    for detail_df in [
        detail_marge_negative(df_erp),
        detail_coherence_stock(df_erp),
        detail_sku_non_numerique(df_web_clean),
    ]:
        if not detail_df.empty:
            detail_df = detail_df.copy()
            detail_df["priorite"] = detail_df.apply(determiner_priorite, axis=1)
            detail_df["action_recommandee"] = detail_df["priorite"].map(ACTION_RECOMMANDEE)
            morceaux.append(detail_df)

    # Règle 10 (cross-dataframe, cf test_pandera.py)
    ok, orphelins = verifier_correspondance_erp_liaison(df_erp, df_liaison)
    if not ok:
        lignes = pd.DataFrame(
            {
                "source": "correspondance erp/liaison",
                "index": orphelins["product_id"],
                "column": "product_id",
                "check": "correspondance_totale",
                "failure_case": orphelins["_merge"].astype(str),
                "priorite": "🟠 Majeure",
                "action_recommandee": ACTION_RECOMMANDEE["🟠 Majeure"],
            }
        )
        morceaux.append(lignes)

    rapport = pd.concat(morceaux, ignore_index=True)

    # Regroupement par (source, ligne) : une même ligne peut violer plusieurs
    # règles (ex : prix négatif ET marge négative sur le même produit). Pour
    # le service achats, mieux vaut UNE ligne par produit avec tous ses
    # problèmes listés, plutôt que le même produit répété plusieurs fois.
    ordre_priorite = {"🔴 Critique": 0, "🟠 Majeure": 1, "🟡 Mineure": 2}
    rapport["ordre_priorite"] = rapport["priorite"].map(ordre_priorite)

    rapport = (
        rapport.groupby(["source", "index"], as_index=False)
        .agg(
            colonnes=("column", lambda s: " + ".join(sorted(set(s.dropna())))),
            regles=("check", lambda s: " ; ".join(sorted(set(s.dropna())))),
            valeurs=("failure_case", lambda s: " | ".join(str(v) for v in s)),
            priorite=("priorite", lambda s: sorted(s, key=lambda p: ordre_priorite[p])[0]),
            nb_regles_violees=("check", "nunique"),
        )
    )
    rapport["action_recommandee"] = rapport["priorite"].map(ACTION_RECOMMANDEE)

    # Tri par priorité (critique en premier) pour que le service achats voie
    # immédiatement ce qui compte le plus, sans avoir à trier lui-même
    rapport["ordre_tri"] = rapport["priorite"].map(ordre_priorite)
    rapport = rapport.sort_values("ordre_tri").drop(columns="ordre_tri").reset_index(drop=True)

    rapport = rapport.rename(
        columns={
            "source": "Fichier / table",
            "index": "Ligne concernée",
            "colonnes": "Colonne(s) concernée(s)",
            "regles": "Règle(s) violée(s)",
            "valeurs": "Valeur(s) observée(s)",
            "priorite": "Priorité",
            "nb_regles_violees": "Nb règles violées",
            "action_recommandee": "Action recommandée",
        }
    )
    return rapport


def exporter_excel(rapport, chemin_sortie):
    """Exporte le rapport en Excel avec mise en forme lisible (Décision 9)."""
    with pd.ExcelWriter(chemin_sortie, engine="openpyxl") as writer:
        rapport.to_excel(writer, sheet_name="Anomalies", index=False)
        feuille = writer.sheets["Anomalies"]

        # En-tête : police pro, gras, fond gris clair
        for col_idx, colonne in enumerate(rapport.columns, start=1):
            cellule = feuille.cell(row=1, column=col_idx)
            cellule.font = Font(name="Arial", bold=True, size=11)
            cellule.fill = PatternFill("solid", fgColor="D9D9D9")
            cellule.alignment = Alignment(horizontal="center", vertical="center")

        # Couleur de fond par ligne selon la priorité (lecture immédiate)
        couleurs_priorite = {
            "🔴 Critique": "F8CBAD",
            "🟠 Majeure": "FFE699",
            "🟡 Mineure": "FFF2CC",
        }
        col_priorite = rapport.columns.get_loc("Priorité") + 1
        for row_idx in range(2, len(rapport) + 2):
            valeur_priorite = feuille.cell(row=row_idx, column=col_priorite).value
            couleur = couleurs_priorite.get(valeur_priorite)
            if couleur:
                for col_idx in range(1, len(rapport.columns) + 1):
                    feuille.cell(row=row_idx, column=col_idx).font = Font(name="Arial", size=10)
                    feuille.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        "solid", fgColor=couleur
                    )

        # Largeur de colonnes lisible
        largeurs = [22, 15, 18, 30, 40, 14, 55]
        for i, largeur in enumerate(largeurs, start=1):
            feuille.column_dimensions[get_column_letter(i)].width = largeur

        feuille.freeze_panes = "A2"  # en-tête toujours visible au scroll


if __name__ == "__main__":
    import os

    rapport = generer_rapport()
    print(f"{len(rapport)} anomalie(s) détectée(s) au total")
    print(rapport["Priorité"].value_counts())

    _racine = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dossier_sortie = os.path.join(_racine, "outputs")
    os.makedirs(dossier_sortie, exist_ok=True)
    chemin_sortie = os.path.join(dossier_sortie, "rapport_anomalies_bottleneck.xlsx")
    exporter_excel(rapport, chemin_sortie)
    print(f"\nRapport exporté : {chemin_sortie}")
